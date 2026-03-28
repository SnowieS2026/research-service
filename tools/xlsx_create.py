"""
xlsx_create.py -- Create Excel (.xlsx) spreadsheets from structured input.
Usage:
  python xlsx_create.py --output "out.xlsx" --sheets "Sheet1,Sheet2" --data "Sheet1,A1:C3,Header|Col A,Col B,Col C"

Sheet spec format: --sheets "Sheet1,Sheet2,Sheet3"

Data spec format: "SheetName,range,type,data"
  Example: "Sheet1,A1:C3,header,Header A,Header B,Header C|Row1,R1C1,R1C2|Row2,R2C1,R2C2"
  Example: "Sheet1,A1:B5,data,Name,Score|Alice,95|Bob,87"

Type: header (first row bold + bg color) | data (normal) | formula

Cell types in data:
  number:N  -> numeric value N
  formula:EXPR -> Excel formula e.g. SUM(B1:B3)
  date:YYYY-MM-DD -> date
  currency:N -> GBP currency
  percentage:N -> percentage format

Styles:
  --style "Sheet1!A1:C3,header=true,bold=true,bg=4472C4,fg=FFFFFF"
  --style "Sheet1!A1:B5,bg=F2F2F2,border=true,align=center"

Options:
  --output             Output .xlsx path (required)
  --sheets             Comma-separated sheet names
  --data               Sheet,range,type,data (repeatable)
  --style              Cell/style options (repeatable)
  --freeze             Freeze panes: "Sheet1!A2" or "Sheet1!B2" etc.
  --chart              Add chart: "Sheet1!A1:B5,bar,Title"
"""

import argparse
import re
import sys
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# Colour map
COLOURS = {
    'header_blue': '4472C4',
    'header_dark': '1F3864',
    'alt_row': 'EBF3FB',
    'accent': '00D4A1',
    'light_gray': 'F2F2F2',
    'white': 'FFFFFF',
    'black': '000000',
    'red': 'FF0000',
    'green': '00B050',
    'yellow': 'FFFF00',
}


def parse_cell_value(value_str):
    """
    Parse a cell value string into (type, python_value, display_value).
    Types: number, formula, text, date, currency, percentage, boolean
    """
    value_str = value_str.strip()
    if not value_str or value_str == '':
        return 'text', '', ''

    # Boolean
    if value_str.lower() in ('true', 'false'):
        return 'boolean', value_str.lower() == 'true', value_str

    # Formula
    if value_str.startswith('formula:') or value_str.startswith('='):
        formula = value_str.lstrip('=').lstrip('formula:')
        return 'formula', formula, ''

    # Currency: currency:100.50
    if value_str.startswith('currency:'):
        num_str = value_str[9:]
        try:
            num = float(num_str)
            return 'currency', num, f'£{num:,.2f}'
        except ValueError:
            return 'text', value_str, value_str

    # Percentage: percentage:0.15 or percentage:15
    if value_str.startswith('percentage:'):
        num_str = value_str[11:]
        try:
            if '.' in num_str:
                num = float(num_str)
            else:
                num = float(num_str) / 100
            return 'percentage', num, f'{num*100:.1f}%'
        except ValueError:
            return 'text', value_str, value_str

    # Date: date:2024-01-15
    if value_str.startswith('date:'):
        date_str = value_str[5:]
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            return 'date', d, date_str
        except ValueError:
            return 'text', date_str, date_str

    # Number
    try:
        if '.' in value_str:
            return 'number', float(value_str), value_str
        else:
            return 'number', int(value_str), value_str
    except ValueError:
        pass

    return 'text', value_str, value_str


def apply_style(cell, style_str):
    """Apply style string to a cell."""
    if not style_str:
        return
    parts = style_str.split(',')
    for part in parts:
        part = part.strip()
        if part == 'bold' or part == 'bold=true':
            cell.font = Font(bold=True)
        elif part == 'italic':
            cell.font = Font(italic=True)
        elif part.startswith('bg=') or part.startswith('bg ='):
            colour = part.split('=')[1].strip().upper()
            if colour in COLOURS:
                colour = COLOURS[colour]
            cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type='solid')
        elif part.startswith('fg=') or part.startswith('fg ='):
            colour = part.split('=')[1].strip().upper()
            if colour in COLOURS:
                colour = COLOURS[colour]
            cell.font = Font(color=colour)
        elif part == 'wrap' or part == 'wrap=true':
            cell.alignment = Alignment(wrap_text=True)
        elif part == 'center' or part == 'align=center':
            cell.alignment = Alignment(horizontal='center')
        elif part == 'right' or part == 'align=right':
            cell.alignment = Alignment(horizontal='right')
        elif part == 'border':
            thin = Side(style='thin')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        elif part.startswith('size='):
            size = int(part.split('=')[1])
            cell.font = Font(size=size)


def parse_data_spec(spec_str):
    """
    Parse a data spec string.
    Format: "SheetName,range,type,CSV_DATA"
    or: "SheetName,range,type,CSV_DATA|style_spec"
    """
    # Split off style part
    if '|' in spec_str and not spec_str.startswith('style:'):
        parts = spec_str.split('|')
        data_part = parts[0]
        style_part = '|'.join(parts[1:])
    else:
        data_part = spec_str
        style_part = None

    sections = data_part.split(',')
    if len(sections) < 4:
        return None

    sheet_name = sections[0]
    range_str = sections[1]
    data_type = sections[2]
    # The rest is the actual data (may contain commas)
    csv_data = ','.join(sections[3:])

    return {
        'sheet': sheet_name,
        'range': range_str,
        'type': data_type,
        'data': csv_data,
        'style': style_part
    }


def build_workbook(sheets, data_specs, style_specs, freeze_specs):
    """Build the workbook from specifications."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name in sheets:
        ws = wb.create_sheet(title=sheet_name)

    for spec_str in data_specs:
        spec = parse_data_spec(spec_str)
        if not spec:
            print(f"  WARNING: Could not parse spec: {spec_str}")
            continue

        sheet_name = spec['sheet']
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found. Skipping.")
            continue

        ws = wb[sheet_name]
        data_type = spec['type']
        style_str = spec.get('style', '')

        rows = spec['data'].split('|')
        for r_idx, row_str in enumerate(rows):
            cells = row_str.split(',')
            for c_idx, cell_str in enumerate(cells):
                if not cell_str.strip() and cell_str != '0':
                    continue
                cell_type, py_val, display = parse_cell_value(cell_str)
                col_letter = get_column_letter(c_idx + 1)
                cell = ws[f'{col_letter}{r_idx + 1}']

                if cell_type == 'formula':
                    cell.value = py_val
                elif cell_type == 'number':
                    cell.value = py_val
                    cell.number_format = '#,##0.00'
                elif cell_type == 'currency':
                    cell.value = py_val
                    cell.number_format = '£#,##0.00'
                elif cell_type == 'percentage':
                    cell.value = py_val
                    cell.number_format = '0.0%'
                elif cell_type == 'date':
                    cell.value = py_val
                    cell.number_format = 'YYYY-MM-DD'
                elif cell_type == 'boolean':
                    cell.value = py_val
                else:
                    cell.value = cell_str.strip()

                # Apply type-specific styling
                if data_type == 'header' and r_idx == 0:
                    cell.font = Font(bold=True)
                    if 'bg' not in style_str:
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')

                if style_str:
                    apply_style(cell, style_str)

    # Apply freeze panes
    for freeze_spec in freeze_specs:
        if '!' in freeze_spec:
            sheet_part, cell_ref = freeze_spec.split('!')
            if sheet_part in wb.sheetnames:
                wb[sheet_part].freeze_panes = cell_ref

    return wb


def main():
    parser = argparse.ArgumentParser(description='Create Excel spreadsheets')
    parser.add_argument('--output', '-o', required=True, help='Output .xlsx path')
    parser.add_argument('--sheets', '-s', required=True, help='Comma-separated sheet names')
    parser.add_argument('--data', '-d', action='append', help='Sheet,range,type,CSV_DATA (repeatable)')
    parser.add_argument('--style', action='append', help='Style spec: "Sheet!range,styles"')
    parser.add_argument('--freeze', action='append', help='Freeze panes: "Sheet!A2"')
    parser.add_argument('--chart', action='append', help='Add chart: "Sheet!range,type,title"')
    args = parser.parse_args()

    sheets = [s.strip() for s in args.sheets.split(',')]
    data_specs = args.data or []
    style_specs = args.style or []
    freeze_specs = args.freeze or []

    wb = build_workbook(sheets, data_specs, style_specs, freeze_specs)

    # Add charts
    for chart_spec in (args.chart or []):
        parts = [p.strip() for p in chart_spec.split(',')]
        if len(parts) < 3:
            continue
        sheet_range = parts[0]
        chart_type = parts[1]
        title = parts[2]

        if '!' not in sheet_range:
            continue
        sheet_name, cell_range = sheet_range.split('!')

        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Parse range like A1:B5
        if ':' in cell_range:
            start_cell, end_cell = cell_range.split(':')
            # Get dimensions
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            start_col, start_row = coordinate_from_string(start_cell)
            end_col, end_row = coordinate_from_string(end_cell)

            data = Reference(ws, min_col=1, min_row=start_row, max_col=end_col, max_row=end_row)

            chart = BarChart()
            chart.add_data(data, titles_from_data=True)
            chart.title = title
            chart.style = 10
            ws.add_chart(chart, start_cell)

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print(f"Created: {output} ({output.stat().st_size:,} bytes)")


if __name__ == '__main__':
    main()
